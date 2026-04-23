"""
MCTS-based Excel parser for discovering optimal card extraction patterns.

Treats Hellcube parsing as a search problem:
- State: Current row position + extracted field mappings
- Actions: Try different field detection strategies (vertical offsets, grouping patterns)
- Reward: Number of valid cards extracted with complete required fields
- MCTS: Discovers optimal parsing strategy through tree search

This is a specialized MCTS for parsing, separate from the layout optimization MCTS.
"""

import random
import math
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ParsingState:
    """
    Represents current state of parsing exploration.

    Tracks:
    - Current row being examined
    - Field mappings discovered (row_idx -> field_name)
    - Card boundaries detected
    - Extracted card data
    """
    current_row: int = 1
    field_mappings: Dict[int, str] = field(default_factory=dict)  # row -> field_name
    card_groups: List[Tuple[int, int]] = field(default_factory=list)  # (start_row, end_row)
    cards_extracted: int = 0
    is_terminal: bool = False

    def copy(self):
        """Create a copy of this state for tree search."""
        return ParsingState(
            current_row=self.current_row,
            field_mappings=dict(self.field_mappings),
            card_groups=list(self.card_groups),
            cards_extracted=self.cards_extracted,
            is_terminal=self.is_terminal
        )


@dataclass
class ParsingAction:
    """
    Represents a parsing strategy action.

    Actions can be:
    - Skip N rows (detect spacing between card groups)
    - Mark field boundary (identify where cards start/end)
    - Extract card group (parse fields at current position)
    """
    action_type: str  # "skip", "mark_boundary", "extract_card"
    parameter: int = 0  # Number of rows to skip, or card group index

    def __repr__(self):
        return f"{self.action_type}({self.parameter})"


class MCTSParserNode:
    """Node in MCTS tree for parsing strategy exploration."""

    def __init__(self, state: ParsingState, parent=None, action: Optional[ParsingAction] = None):
        self.state = state
        self.parent = parent
        self.action = action  # Action that led to this state

        self.children: List['MCTSParserNode'] = []
        self.untried_actions: List[ParsingAction] = []

        self.visits = 0
        self.total_reward = 0.0

    def is_fully_expanded(self) -> bool:
        """Check if all actions from this node have been tried."""
        return len(self.untried_actions) == 0

    def best_child(self, exploration_weight: float = 1.414) -> 'MCTSParserNode':
        """
        Select best child using UCB1 formula.

        UCB1 = Q(node) + C * sqrt(ln(N_parent) / N_node)
        where C = exploration_weight (√2 ≈ 1.414)
        """
        best_score = -float('inf')
        best_child = None

        for child in self.children:
            if child.visits == 0:
                # Unvisited nodes have infinite priority
                return child

            # Average reward
            q_value = child.total_reward / child.visits

            # Exploration bonus
            exploration = exploration_weight * math.sqrt(math.log(self.visits) / child.visits)

            ucb1_score = q_value + exploration

            if ucb1_score > best_score:
                best_score = ucb1_score
                best_child = child

        return best_child


class MCTSParser:
    """
    MCTS-based parser for discovering optimal card extraction strategy.

    Uses Monte Carlo Tree Search to explore different parsing strategies
    and find the one that extracts the most valid cards.
    """

    FIELD_LABELS = ["name", "pic", "types", "text", "flavor", "stats", "author"]

    def __init__(self, sheet: Worksheet, max_rollouts: int = 500, debug: bool = False):
        """
        Initialize MCTS parser.

        Args:
            sheet: Excel worksheet to parse
            max_rollouts: Maximum MCTS iterations (default 500 for thorough search)
            debug: If True, print debug information during search
        """
        self.sheet = sheet
        self.max_rollouts = max_rollouts
        self.max_row = sheet.max_row
        self.debug = debug
        self.best_cards_found = 0  # Track best result during search
        self.best_state_found = None  # Track actual best state from simulations

    def find_optimal_parsing_strategy(self) -> ParsingState:
        """
        Use MCTS to discover optimal parsing strategy.

        Returns:
            Best parsing state found (with field mappings and card boundaries)
        """
        # Initialize root state
        root_state = ParsingState(current_row=1)
        root_node = MCTSParserNode(root_state)

        # Initialize possible actions from root
        root_node.untried_actions = self._generate_actions(root_state)

        # MCTS main loop
        for rollout in range(self.max_rollouts):
            # 1. Selection: traverse tree to leaf using UCB1
            node = self._select(root_node)

            # 2. Expansion: add one new child
            if not node.state.is_terminal and not node.is_fully_expanded():
                node = self._expand(node)

            # 3. Simulation: random rollout from new node
            reward, final_state = self._simulate(node.state)

            # 4. Backpropagation: update ancestors
            self._backpropagate(node, reward)

            # Track best result AND best state
            if reward > self.best_cards_found:
                self.best_cards_found = int(reward)
                self.best_state_found = final_state  # Save the actual state that achieved this reward!
                if self.debug:
                    print(f"  [Rollout {rollout+1}/{self.max_rollouts}] New best: {self.best_cards_found} cards, {len(final_state.card_groups)} groups")

            # Progress reporting every 100 rollouts
            if self.debug and (rollout + 1) % 100 == 0:
                print(f"  [Progress] Rollout {rollout+1}/{self.max_rollouts}, Best so far: {self.best_cards_found} cards")

        # Return the actual best state found during simulations
        if self.best_state_found:
            if self.debug:
                print(f"\n✅ MCTS complete! Best simulation state has {self.best_state_found.cards_extracted} cards")
                print(f"   Card groups found: {len(self.best_state_found.card_groups)}")
            return self.best_state_found
        elif root_node.children:
            # Fallback: return best child state
            best_child = max(root_node.children, key=lambda c: c.total_reward)
            if self.debug:
                print(f"\n✅ MCTS complete! Best child has {best_child.state.cards_extracted} cards in state")
            return best_child.state
        else:
            if self.debug:
                print("⚠️  No children expanded from root!")
            return root_state

    def _select(self, node: MCTSParserNode) -> MCTSParserNode:
        """
        Selection phase: traverse tree using UCB1 until reaching expandable node.
        """
        while node.is_fully_expanded() and node.children and not node.state.is_terminal:
            node = node.best_child()
        return node

    def _expand(self, node: MCTSParserNode) -> MCTSParserNode:
        """
        Expansion phase: add one new child for an untried action.
        """
        if not node.untried_actions:
            return node

        # Pop one untried action
        action = node.untried_actions.pop()

        # Apply action to create new state
        new_state = self._apply_action(node.state, action)

        # Create child node
        child_node = MCTSParserNode(new_state, parent=node, action=action)
        child_node.untried_actions = self._generate_actions(new_state)

        node.children.append(child_node)
        return child_node

    def _simulate(self, state: ParsingState) -> tuple[float, ParsingState]:
        """
        Simulation phase: random rollout to estimate state value.

        Returns:
            Tuple of (reward, final_state)
            Reward = number of valid cards extracted with all required fields.
        """
        current_state = state.copy()

        # Random rollout: try random actions until terminal
        depth = 0
        max_depth = 200  # Increased to explore more card groups (Hellcube has ~50 groups)

        while not current_state.is_terminal and depth < max_depth:
            actions = self._generate_actions(current_state)
            if not actions:
                break

            # Random action selection
            action = random.choice(actions)
            current_state = self._apply_action(current_state, action)
            depth += 1

        # Evaluate final state: reward = cards extracted with required fields
        reward = float(current_state.cards_extracted)
        return reward, current_state

    def _backpropagate(self, node: MCTSParserNode, reward: float):
        """
        Backpropagation phase: update visit counts and rewards up the tree.
        """
        while node is not None:
            node.visits += 1
            node.total_reward += reward
            node = node.parent

    def _generate_actions(self, state: ParsingState) -> List[ParsingAction]:
        """
        Generate possible parsing actions from current state.

        Actions:
        - Skip 1-5 rows (detect spacing between card groups)
        - Extract card at current position (if field pattern detected)
        - GREEDY: Jump to next known name row (informed search)
        """
        if state.is_terminal or state.current_row >= self.max_row:
            return []

        actions = []

        # Action 1: Extract card if we're at a name row
        if self._looks_like_name_row(state.current_row):
            actions.append(ParsingAction("extract_card", state.current_row))
            return actions  # Prioritize extraction!

        # Action 2: GREEDY - Find next name row within next 20 rows
        for look_ahead in range(1, 21):
            next_row = state.current_row + look_ahead
            if next_row >= self.max_row:
                break
            if self._looks_like_name_row(next_row):
                # Found a name row! Add action to skip to it
                actions.append(ParsingAction("skip", look_ahead))
                break  # Only add first name row found

        # Action 3: Fallback - Random exploration if no name rows nearby
        if not actions:
            for skip in [5, 10, 15]:
                if state.current_row + skip < self.max_row:
                    actions.append(ParsingAction("skip", skip))

        return actions

    def _apply_action(self, state: ParsingState, action: ParsingAction) -> ParsingState:
        """
        Apply parsing action to state, returning new state.
        """
        new_state = state.copy()

        if action.action_type == "skip":
            # Skip rows to find next card group
            new_state.current_row += action.parameter

        elif action.action_type == "extract_card":
            # Try to extract card starting at this row
            card_data = self._try_extract_card(action.parameter)

            if card_data:
                # Valid card found!
                new_state.cards_extracted += 1
                new_state.card_groups.append((action.parameter, action.parameter + 10))
                # Move just 1 row forward to allow greedy search to find next card group
                # (some groups are 9 rows, some are 10+ rows)
                new_state.current_row = action.parameter + 1
            else:
                # Failed extraction, skip ahead
                new_state.current_row += 1

        # Terminal condition: reached end of sheet or extracted ALL possible cards
        # (Hellcube has ~200 cards, so set high limit)
        if new_state.current_row >= self.max_row or new_state.cards_extracted >= 200:
            new_state.is_terminal = True

        return new_state

    def _looks_like_name_row(self, row_idx: int) -> bool:
        """
        Check if row looks like it contains card names.

        Heuristic: Column A contains "name" label (case-insensitive).
        """
        if row_idx > self.max_row:
            return False

        label_cell = self.sheet.cell(row_idx, 1)
        label = str(label_cell.value).strip().lower() if label_cell.value else ""

        # Debug: print when we find name rows
        if label == "name":
            col_c_val = self.sheet.cell(row_idx, 3).value
            # print(f"  [DEBUG] Found 'name' at row {row_idx}, Column C = {col_c_val}")

        return label == "name"

    def _try_extract_card(self, start_row: int) -> Optional[Dict]:
        """
        Try to extract a card starting at given row.

        Returns:
            Card data dict if successful, None if incomplete
        """
        # Simplified extraction for MCTS simulation
        # Check if we can find name and types fields (required per FR-013)

        card_data = {}

        for row_offset in range(10):  # Check next 10 rows for fields
            row_idx = start_row + row_offset
            if row_idx > self.max_row:
                break

            label_cell = self.sheet.cell(row_idx, 1)
            label = str(label_cell.value).strip().lower() if label_cell.value else ""

            if label == "name":
                # Check column C (3) for card name
                value_cell = self.sheet.cell(row_idx, 3)
                if value_cell.value:
                    card_data["name"] = str(value_cell.value)

            elif label == "types":
                # Check column C for types
                value_cell = self.sheet.cell(row_idx, 3)
                if value_cell.value:
                    card_data["types"] = str(value_cell.value)

        # Valid if has both required fields
        if "name" in card_data and "types" in card_data:
            return card_data
        else:
            return None
