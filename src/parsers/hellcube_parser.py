"""
Hellcube Excel parser with dynamic adjacency detection.

Parses unstructured Hellcube AJ.xlsx spreadsheet using semantic pattern
recognition to extract MTG card data.

Implements:
- FR-001: Semantic pattern detection with dynamic adjacency
- FR-003: Combine multiple "text" rows into abilities list
- FR-004: Extract power/toughness from Stats field (datetime format)
- FR-004a: Infer primary card type from Types field
- FR-004b: Detect legendary status from "Legendary" keyword
- FR-013: Validate required fields (name, type)
"""

import re
from datetime import datetime
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.models.card import Card
from src.parsers.mana_cost_parser import extract_mana_cost_from_name, parse_mana_cost
from src.parsers.color_inference import infer_color_from_mana
from src.parsers.mcts_parser import MCTSParser


class HellcubeExcelParser:
    """
    Parser for Hellcube AJ.xlsx spreadsheet format.

    Uses dynamic adjacency detection to find card data in columns marked by
    "AJ" markers, with field labels in column A and values in neighboring cells.
    """

    # Field labels expected in column A
    FIELD_LABELS = {
        "name": "name",
        "pic": "pic",
        "types": "Types",
        "text": "text",
        "flavor": "flavor",
        "stats": "Stats",
        "author": "Author",
    }

    def __init__(self, file_path: str):
        """
        Initialize parser with Excel file path.

        Args:
            file_path: Path to Hellcube AJ.xlsx file
        """
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet: Worksheet = self.workbook.active

    def parse_all_cards(self, use_mcts: bool = True) -> List[Card]:
        """
        Parse all cards from the Hellcube spreadsheet.

        Args:
            use_mcts: If True, use MCTS to discover optimal parsing strategy.
                     If False, use simplified first-group-only parser.

        Returns:
            List of Card objects extracted from spreadsheet

        Implements FR-001: Dynamic adjacency detection across columns 2-9
        """
        if use_mcts:
            # MCTS-based parsing: discover optimal card extraction strategy
            mcts_parser = MCTSParser(self.sheet, max_rollouts=100)
            print("🔍 Running MCTS to discover optimal parsing strategy...")
            optimal_state = mcts_parser.find_optimal_parsing_strategy()
            print(f"✅ MCTS found {len(optimal_state.card_groups)} card groups!")

            # Use discovered card groups to extract all cards
            return self._parse_with_discovered_groups(optimal_state.card_groups)
        else:
            # Simplified parsing: first card group only (Milestone 1)
            card_columns = self._find_card_columns()
            cards = []
            for col_idx in card_columns:
                try:
                    card = self._parse_card_column(col_idx)
                    if card:
                        cards.append(card)
                except Exception as e:
                    print(f"Warning: Failed to parse card in column {col_idx}: {e}")
                    continue
            return cards

    def _parse_with_discovered_groups(self, card_groups: List[Tuple[int, int]]) -> List[Card]:
        """
        Parse cards using MCTS-discovered card group boundaries.

        Args:
            card_groups: List of (start_row, end_row) tuples for each card group

        Returns:
            List of all cards extracted
        """
        card_columns = self._find_card_columns()
        all_cards = []

        for start_row, end_row in card_groups:
            for col_idx in card_columns:
                try:
                    card = self._parse_card_in_range(col_idx, start_row, end_row)
                    if card:
                        all_cards.append(card)
                except Exception as e:
                    print(f"Warning: Failed to parse card at rows {start_row}-{end_row}, col {col_idx}: {e}")
                    continue

        return all_cards

    def _parse_card_in_range(self, col_idx: int, start_row: int, end_row: int) -> Optional[Card]:
        """
        Parse a card from specific row range.

        Args:
            col_idx: Column containing card data
            start_row: Start of card group
            end_row: End of card group

        Returns:
            Card object or None
        """
        card_data = {}
        text_rows = []
        first_name_row_seen = False

        # Scan only within specified row range
        for row_idx in range(start_row, min(end_row + 1, self.sheet.max_row + 1)):
            label_cell = self.sheet.cell(row_idx, 1)
            label = str(label_cell.value).strip().lower() if label_cell.value else ""

            if label == "name":
                # Stop if we hit a second "name" row (next card group boundary)
                if first_name_row_seen:
                    break  # Hit next card group, stop here!
                first_name_row_seen = True
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["name"] = str(value_cell.value).strip() if value_cell.value else ""
            elif label == "pic":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["pic"] = str(value_cell.value).strip() if value_cell.value else None
            elif label == "types":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["types"] = str(value_cell.value).strip() if value_cell.value else ""
            elif label == "text":
                value_cell = self.sheet.cell(row_idx, col_idx)
                text_value = str(value_cell.value).strip() if value_cell.value else ""
                if text_value:
                    text_rows.append(text_value)
            elif label == "flavor":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["flavor"] = str(value_cell.value).strip() if value_cell.value else None
            elif label == "stats":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["stats"] = value_cell.value
            elif label == "author":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["author"] = str(value_cell.value).strip() if value_cell.value else None

        # Validate required fields
        if not card_data.get("name") or not card_data.get("types"):
            return None

        # Parse card attributes
        clean_name, mana_cost_str = extract_mana_cost_from_name(card_data["name"])
        mana_counts = parse_mana_cost(card_data["name"])
        color = infer_color_from_mana(mana_counts)
        primary_type, legendary, subtypes = self._parse_types_field(card_data["types"])
        power_toughness = self._parse_stats_field(card_data.get("stats"))

        return Card(
            name=clean_name,
            type=primary_type,
            mana_cost=mana_cost_str if mana_cost_str else None,
            color=color,
            legendary=legendary,
            subtypes=subtypes,
            abilities=text_rows,
            flavor=card_data.get("flavor"),
            power_toughness=power_toughness,
            author=card_data.get("author"),
            artwork_url=card_data.get("pic") if card_data.get("pic") else None,
        )

    def _find_card_columns(self) -> List[int]:
        """
        Find card data columns by locating "AJ" markers.

        Scans columns 2-9 (B-I) for cells containing "AJ" marker.

        Returns:
            List of column indices containing card data

        Implements FR-001: Discover cards in columns C, E, G, I with
        dynamic detection (not hardcoded)
        """
        card_columns = []

        # Scan columns 2-9 (FR-001: columns 2-9 range)
        for col_idx in range(2, 10):  # Columns B through I
            for row_idx in range(1, min(20, self.sheet.max_row + 1)):
                cell_value = self.sheet.cell(row_idx, col_idx).value
                if cell_value and str(cell_value).strip() == "AJ":
                    card_columns.append(col_idx)
                    break  # Found marker in this column

        return card_columns

    def _parse_card_column(self, col_idx: int) -> Optional[Card]:
        """
        Parse a single card from a column.

        Args:
            col_idx: Column index containing card data

        Returns:
            Card object or None if parsing fails

        Implements adjacency detection: field labels in column A,
        values in neighboring column.

        SIMPLIFIED VERSION (Milestone 1): Parse first card group only.
        TODO: Implement full vertical card group detection for all ~200 cards.
        """
        card_data = {}
        text_rows = []  # Collect multiple "text" rows (FR-003)

        # SIMPLIFIED: Only scan first 30 rows for first card group
        # (Full implementation will detect multiple card groups vertically)
        for row_idx in range(1, min(30, self.sheet.max_row + 1)):
            label_cell = self.sheet.cell(row_idx, 1)  # Column A
            label = str(label_cell.value).strip().lower() if label_cell.value else ""

            # Check if this is a recognized field label
            if label == "name":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["name"] = str(value_cell.value).strip() if value_cell.value else ""
            elif label == "pic":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["pic"] = str(value_cell.value).strip() if value_cell.value else None
            elif label == "types":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["types"] = str(value_cell.value).strip() if value_cell.value else ""
            elif label == "text":
                # Multiple "text" rows - collect all (FR-003)
                value_cell = self.sheet.cell(row_idx, col_idx)
                text_value = str(value_cell.value).strip() if value_cell.value else ""
                if text_value:
                    text_rows.append(text_value)
            elif label == "flavor":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["flavor"] = str(value_cell.value).strip() if value_cell.value else None
            elif label == "stats":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["stats"] = value_cell.value  # Keep as datetime or string
            elif label == "author":
                value_cell = self.sheet.cell(row_idx, col_idx)
                card_data["author"] = str(value_cell.value).strip() if value_cell.value else None

        # Validate required fields (FR-013)
        if not card_data.get("name") or not card_data.get("types"):
            return None  # Skip incomplete cards

        # Parse card name and mana cost (FR-002)
        clean_name, mana_cost_str = extract_mana_cost_from_name(card_data["name"])
        mana_counts = parse_mana_cost(card_data["name"])
        color = infer_color_from_mana(mana_counts)

        # Parse Types field (FR-004a, FR-004b)
        primary_type, legendary, subtypes = self._parse_types_field(card_data["types"])

        # Parse Stats field (FR-004)
        power_toughness = self._parse_stats_field(card_data.get("stats"))

        # Build Card object
        return Card(
            name=clean_name,
            type=primary_type,
            mana_cost=mana_cost_str if mana_cost_str else None,
            color=color,
            legendary=legendary,
            subtypes=subtypes,
            abilities=text_rows,  # FR-003: multiple text rows combined
            flavor=card_data.get("flavor"),
            power_toughness=power_toughness,
            author=card_data.get("author"),
            artwork_url=card_data.get("pic") if card_data.get("pic") else None,
        )

    def _parse_types_field(self, types_str: str) -> tuple[str, bool, List[str]]:
        """
        Parse Types field into primary type, legendary status, and subtypes.

        Args:
            types_str: Types field value like "Creature- Human, Batman"

        Returns:
            Tuple of (primary_type, legendary, subtypes)

        Implements:
        - FR-004a: Extract primary type from first word after dash
        - FR-004b: Detect "Legendary" keyword
        """
        if not types_str:
            return "Unknown", False, []

        # Check for Legendary keyword (FR-004b)
        legendary = "legendary" in types_str.lower()

        # Remove "Legendary" prefix if present
        clean_types = re.sub(r"legendary\s+", "", types_str, flags=re.IGNORECASE).strip()

        # Split by dash to separate primary type from subtypes
        if "-" in clean_types:
            parts = clean_types.split("-", 1)
            primary_type = parts[0].strip()
            subtypes_str = parts[1].strip() if len(parts) > 1 else ""
            subtypes = [s.strip() for s in subtypes_str.split(",") if s.strip()]
        else:
            # No dash - entire string is primary type
            primary_type = clean_types.strip()
            subtypes = []

        return primary_type, legendary, subtypes

    def _parse_stats_field(self, stats_value) -> Optional[str]:
        """
        Parse Stats field into power/toughness string.

        Args:
            stats_value: Stats field value (datetime or string)

        Returns:
            Power/toughness string like "2/4" or None

        Implements FR-004: Handle datetime format "2025-02-04" → "2/4"
        """
        if stats_value is None:
            return None

        # If it's a datetime object, extract month/day as power/toughness
        if isinstance(stats_value, datetime):
            power = stats_value.month
            toughness = stats_value.day
            return f"{power}/{toughness}"

        # If it's already a string like "2/4", return as-is
        stats_str = str(stats_value).strip()
        if "/" in stats_str:
            return stats_str

        # Try to parse as datetime string
        try:
            dt = datetime.fromisoformat(stats_str.split()[0])  # Handle "2025-02-04 00:00:00"
            return f"{dt.month}/{dt.day}"
        except (ValueError, AttributeError):
            return None
